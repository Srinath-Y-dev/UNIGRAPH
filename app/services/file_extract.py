from __future__ import annotations
import csv, hashlib
from pathlib import Path
from typing import List,Dict
from pypdf import PdfReader
from openpyxl import load_workbook
try:
    from PIL import Image
except Exception:
    Image=None

def extract_file(path:Path)->List[Dict[str,str]]:
    ext=path.suffix.lower()
    if ext=='.pdf':
        reader=PdfReader(str(path)); out=[]
        for i,p in enumerate(reader.pages,1):
            text=p.extract_text() or ''
            if text.strip(): out.append({'name':f'{path.name} · page {i}','type':'manufacturer_datasheet','text':text})
        return out or [{'name':path.name,'type':'manufacturer_datasheet','text':'Scanned PDF with no embedded text. OCR/VLM connector required for page content.'}]
    if ext in {'.txt','.md'}: return [{'name':path.name,'type':'customer_catalog','text':path.read_text(encoding='utf-8',errors='ignore')}]
    if ext=='.csv':
        rows=[]
        with path.open('r',encoding='utf-8-sig',errors='ignore',newline='') as f:
            for row in csv.DictReader(f): rows.append(' | '.join(f'{k}: {v}' for k,v in row.items() if v not in (None,'')))
        return [{'name':path.name,'type':'customer_catalog','text':'\n'.join(rows)}]
    if ext in {'.xlsx','.xlsm'}:
        wb=load_workbook(path,read_only=True,data_only=True); parts=[]
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                vals=[str(v) for v in row if v is not None]
                if vals: parts.append(f'[{ws.title}] '+' | '.join(vals))
        return [{'name':path.name,'type':'customer_catalog','text':'\n'.join(parts)}]
    if ext in {'.png','.jpg','.jpeg','.webp'}:
        h=hashlib.sha256(path.read_bytes()).hexdigest()[:16]; meta=f'Image asset: {path.name}; SHA256: {h}'
        if Image:
            try:
                with Image.open(path) as im: meta+=f'; width: {im.width}px; height: {im.height}px; format: {im.format}; mode: {im.mode}'
            except Exception: pass
        return [{'name':path.name,'type':'digital_asset','text':meta+'; visual attributes require configured VLM connector.'}]
    return [{'name':path.name,'type':'customer_catalog','text':f'Unsupported binary asset retained: {path.name}'}]
